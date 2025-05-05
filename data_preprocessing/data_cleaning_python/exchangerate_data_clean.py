import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment
import logging
import re

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def fetch_webpage(url, headers):
    """Fetch webpage content with error handling."""
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        logger.info(f"Webpage {url} fetched successfully")
        return response.text
    except requests.RequestException as e:
        logger.error(f"Failed to fetch webpage {url}: {e}")
        return None


def parse_exchange_rates(html_content, currency="USD"):
    """Parse exchange rate data from HTML tables for February to December."""
    try:
        soup = BeautifulSoup(html_content, 'html.parser')

        # Find all tables
        tables = soup.find_all('table', id='hist')
        if not tables:
            logger.error("No tables with id='hist' found in HTML")
            return None, None

        logger.debug(f"Found {len(tables)} tables with id='hist'")

        dates = []
        rates = []

        # Currency-specific rate indicators
        rate_indicators = {
            "USD": "$1 USD =",
            "CNY": "¥1 CNY =",
            "MOP": "$1 MOP =",
            "TWD": "NT$1 TWD =",
            "KRW": "₩1 KRW ="
        }
        rate_indicator = rate_indicators[currency]

        # Process each table
        for i, table in enumerate(tables):
            # Check if table contains exchange rate data
            if rate_indicator not in table.text:
                logger.debug(f"Table {i} does not contain '{rate_indicator}', skipping")
                continue

            # Log first few rows for debugging
            rows = table.find_all('tr')[:3]
            for j, row in enumerate(rows):
                cells = row.find_all('td')
                logger.debug(f"Table {i} Row {j} cells: {[cell.text.strip() for cell in cells]}")

            # Check if table is for January (skip it)
            first_row = table.find('tr')
            if first_row:
                first_cell = first_row.find('td')
                if first_cell and 'January' in first_cell.text:
                    logger.debug(f"Table {i} is for January, skipping")
                    continue

            logger.debug(f"Processing Table {i} as {currency}/HKD exchange rate table")

            # Process table rows
            for row in table.find_all('tr')[1:]:  # Skip header row
                cells = row.find_all('td')
                if len(cells) >= 3:  # Expect 3 columns: Date, Rate, Link
                    # Log raw cell content
                    logger.debug(f"Processing row cells: {[cell.text.strip() for cell in cells]}")

                    # Extract date (e.g., "Friday  1 February 2019")
                    date_text = cells[0].text.strip()
                    try:
                        # Extract date part (last 3 parts: day month year)
                        date_parts = date_text.split()
                        if len(date_parts) >= 3:
                            date_str = ' '.join(date_parts[-3:])
                            date = datetime.strptime(date_str, '%d %B %Y').strftime('%m/%d/%Y')
                        else:
                            logger.warning(f"Invalid date format: {date_text}")
                            continue
                    except ValueError as e:
                        logger.warning(f"Skipping invalid date: {date_text} ({e})")
                        continue

                    # Extract rate (e.g., "$1 USD = $7.8466", "¥1 CNY = $1.1388", etc.)
                    rate_text = cells[1].text.strip()
                    try:
                        # Use regex to extract the rate (e.g., 7.8466, 1.1388, 0.9756, 0.2547, 0.0066)
                        rate_match = re.search(r'\$(\d+\.\d*)', rate_text)
                        if rate_match:
                            rate = float(rate_match.group(1))
                            dates.append(date)
                            rates.append(rate)
                        else:
                            logger.warning(f"No numeric rate found in: {rate_text}")
                            continue
                    except (ValueError, IndexError) as e:
                        logger.warning(f"Skipping invalid rate: {rate_text} ({e})")
                        continue
                else:
                    logger.debug(f"Skipping row with insufficient cells: {len(cells)}")

        if not dates:
            logger.error("No valid data extracted from tables")
            return None, None

        logger.info(f"Parsed {len(dates)} {currency}/HKD exchange rate entries")
        return dates, rates
    except Exception as e:
        logger.error(f"Error parsing HTML: {e}")
        return None, None


def create_dataframe(dates, rates, column_name):
    """Create a pandas DataFrame from parsed data."""
    if dates and rates and len(dates) == len(rates):
        df = pd.DataFrame({
            'Date': dates,
            column_name: rates
        })
        logger.info(f"{column_name} DataFrame created successfully")
        return df
    logger.error(f"Invalid or empty data for {column_name} DataFrame")
    return None


def format_excel(file_path, sheet_name):
    """Format the Excel file with styles and adjusted column widths."""
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # Define styles
        date_style = NamedStyle(name='date', number_format='MM/DD/YYYY')
        rate_style = NamedStyle(name='rate', number_format='0.0000')
        header_font = Font(bold=True)
        header_align = Alignment(horizontal='center')

        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align

        # Apply styles to data
        for row in ws['A2:A' + str(ws.max_row)]:
            for cell in row:
                cell.style = date_style
        for col in ['B', 'C', 'D', 'E', 'F']:  # Columns for rates
            for row in ws[f'{col}2:{col}' + str(ws.max_row)]:
                for cell in row:
                    cell.style = rate_style

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 3

        wb.save(file_path)
        logger.info(f"Excel file formatted and saved: {file_path}")
    except Exception as e:
        logger.error(f"Error formatting Excel file: {e}")


def main():
    """Main function to scrape and export exchange rates to HKD for 2019-2025."""
    currency_data = {
        "USA": {
            "currency": "USD",
            "urls": [
                "https://www.exchangerates.org.uk/USD-HKD-spot-exchange-rates-history-2019.html",
                "https://www.exchangerates.org.uk/USD-HKD-spot-exchange-rates-history-2020.html",
                "https://www.exchangerates.org.uk/USD-HKD-spot-exchange-rates-history-2023.html",
                "https://www.exchangerates.org.uk/USD-HKD-spot-exchange-rates-history-2024.html",
                "https://www.exchangerates.org.uk/USD-HKD-spot-exchange-rates-history-2025.html"
            ]
        },
        "Mainland": {
            "currency": "CNY",
            "urls": [
                "https://www.exchangerates.org.uk/CNY-HKD-spot-exchange-rates-history-2019.html",
                "https://www.exchangerates.org.uk/CNY-HKD-spot-exchange-rates-history-2020.html",
                "https://www.exchangerates.org.uk/CNY-HKD-spot-exchange-rates-history-2023.html",
                "https://www.exchangerates.org.uk/CNY-HKD-spot-exchange-rates-history-2024.html",
                "https://www.exchangerates.org.uk/CNY-HKD-spot-exchange-rates-history-2025.html"
            ]
        },
        "Macau SAR": {
            "currency": "MOP",
            "urls": [
                "https://www.exchangerates.org.uk/MOP-HKD-spot-exchange-rates-history-2019.html",
                "https://www.exchangerates.org.uk/MOP-HKD-spot-exchange-rates-history-2020.html",
                "https://www.exchangerates.org.uk/MOP-HKD-spot-exchange-rates-history-2023.html",
                "https://www.exchangerates.org.uk/MOP-HKD-spot-exchange-rates-history-2024.html",
                "https://www.exchangerates.org.uk/MOP-HKD-spot-exchange-rates-history-2025.html"
            ]
        },
        "Taiwan": {
            "currency": "TWD",
            "urls": [
                "https://www.exchangerates.org.uk/TWD-HKD-spot-exchange-rates-history-2019.html",
                "https://www.exchangerates.org.uk/TWD-HKD-spot-exchange-rates-history-2020.html",
                "https://www.exchangerates.org.uk/TWD-HKD-spot-exchange-rates-history-2023.html",
                "https://www.exchangerates.org.uk/TWD-HKD-spot-exchange-rates-history-2024.html",
                "https://www.exchangerates.org.uk/TWD-HKD-spot-exchange-rates-history-2025.html"
            ]
        },
        "South Korea": {
            "currency": "KRW",
            "urls": [
                "https://www.exchangerates.org.uk/KRW-HKD-spot-exchange-rates-history-2019.html",
                "https://www.exchangerates.org.uk/KRW-HKD-spot-exchange-rates-history-2020.html",
                "https://www.exchangerates.org.uk/KRW-HKD-spot-exchange-rates-history-2023.html",
                "https://www.exchangerates.org.uk/KRW-HKD-spot-exchange-rates-history-2024.html",
                "https://www.exchangerates.org.uk/KRW-HKD-spot-exchange-rates-history-2025.html"
            ]
        }
    }

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
        'Referer': 'https://www.exchangerates.org.uk/'
    }

    all_dfs = []

    # Process each currency
    for column_name, data in currency_data.items():
        currency = data["currency"]
        urls = data["urls"]
        dfs = []

        for url in urls:
            html_content = fetch_webpage(url, headers)
            if not html_content:
                continue
            dates, rates = parse_exchange_rates(html_content, currency=currency)
            if not dates or not rates:
                continue
            df = create_dataframe(dates, rates, column_name=column_name)
            if df is not None:
                dfs.append(df)

        if dfs:
            combined = pd.concat(dfs, ignore_index=True)
            combined['Date'] = pd.to_datetime(combined['Date'], format='%m/%d/%Y')
            all_dfs.append(combined)

    # Combine all DataFrames
    if all_dfs:
        # Start with the first DataFrame
        combined_df = all_dfs[0]
        # Merge with others on Date
        for df in all_dfs[1:]:
            combined_df = combined_df.merge(df, on='Date', how='outer')

        # Sort by date and format
        combined_df = combined_df.sort_values('Date').reset_index(drop=True)
        combined_df['Date'] = combined_df['Date'].dt.strftime('%m/%d/%Y')

        # Reorder columns as requested
        column_order = ['Date', 'Mainland', 'Taiwan', 'Macau SAR', 'South Korea', 'USA']
        combined_df = combined_df[column_order]

        # Export to Excel
        output_file = 'processed_exchangerate_data.xlsx'
        sheet_name = 'Exchange_Rates'
        try:
            combined_df.to_excel(output_file, index=False, sheet_name=sheet_name)
            format_excel(output_file, sheet_name)
            logger.info(f"Successfully exported {len(combined_df)} rows to {output_file}")
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
    else:
        logger.error("No data to export")


if __name__ == "__main__":
    main()